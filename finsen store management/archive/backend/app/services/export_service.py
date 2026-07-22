import io
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


BRAND_COLOR = colors.HexColor("#1565C0")
HEADER_COLOR = colors.HexColor("#E3F2FD")


def generate_pdf_report(title: str, headers: list[str], rows: list[list], subtitle: str = "") -> bytes:
    """Generate a professional PDF report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=BRAND_COLOR,
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=12,
    )

    elements = []

    # Title
    elements.append(Paragraph("Finsen Riter Limited", title_style))
    elements.append(Paragraph(title, ParagraphStyle("T2", parent=styles["Heading2"], fontSize=14, alignment=TA_CENTER, spaceAfter=4)))
    if subtitle:
        elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", subtitle_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Table
    table_data = [headers] + rows
    col_width = (landscape(A4)[0] - inch) / max(len(headers), 1)
    col_widths = [col_width] * len(headers)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, HEADER_COLOR]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generate_excel_report(title: str, headers: list[str], rows: list[list], sheet_name: str = "Report") -> bytes:
    """Generate a professional Excel report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=16, color="1565C0")
    normal_align = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill("solid", fgColor="E3F2FD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells(f"A1:{chr(64 + len(headers))}1")
    ws["A1"] = "Finsen Riter Limited — " + title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Generated at row
    ws.merge_cells(f"A2:{chr(64 + len(headers))}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", italic=True, color="888888", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = normal_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = 20

    ws.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
